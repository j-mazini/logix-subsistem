#!/usr/bin/env python3
"""
Importa dados dos ficheiros Excel para o DHL Subsystem:
  - DISCO_RouteData_*.xlsx  → assets/disco.json + assets/disco-data.js (Disco)
  - BA Daily Figures 3.xlsx → OPMS (counts, byRoute) em dhl-embedded-data.js
  - TW 300126.xlsx         → Time Window (% TW Adh DL) por rota em dhl-embedded-data.js

Uso:
  python scripts/import_excel_to_data.py
  python scripts/import_excel_to_data.py [path_ba] [path_tw] [path_disco]
  python scripts/import_excel_to_data.py "" "" /caminho/para/DISCO_RouteData.xlsx

Por defeito procura em ~/Downloads:
  - BA Daily Figures 3.xlsx
  - TW 300126.xlsx
  - DISCO_RouteData_20260202_064654.xlsx
"""

import json
import os
import re
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Requer openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# OPMS
OPMS_CODES = frozenset({"BA", "CA", "CM", "FP", "HN", "NR", "OK", "PU", "RD"})
COL_PUD_SVC = "PUD Svc Area"
COL_PUD_RTE = "PUD Rte"
COL_ACT_CKPT = "Act Ckpt Code"
OPMS_SHEET_NAME = "OPMS - RDT ROADMAP Stop File"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT = os.path.dirname(SCRIPT_DIR)
ASSETS_DIR = os.path.join(REPO_ROOT, "assets")
DEFAULT_OPMS_DATE = "2026-01-30"


def _norm_path(p):
    """Normaliza caminho: expande ~ e retorna None se vazio."""
    if not p or not str(p).strip():
        return None
    return os.path.normpath(os.path.expanduser(str(p).strip()))


def find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        row_str = [str(c).strip() if c is not None else "" for c in row]
        if COL_ACT_CKPT in row_str:
            return i, row_str
        if any("Act Ckpt Code" in (s or "") for s in row):
            return i, row_str
    return None, []


def find_col(header_row, name, fallback_sub=None):
    for i, h in enumerate(header_row):
        if h == name:
            return i
        if fallback_sub and (h or "").strip() and fallback_sub in (h or "").upper():
            return i
    return None


def parse_opms(ws):
    """Retorna (counts, by_route) onde by_route é dict "depot|route" -> { depot, route, counts }."""
    header_row_idx, header_row = find_header_row(ws)
    if not header_row:
        return {}, {}

    idx_depot = find_col(header_row, COL_PUD_SVC, "PUD SVC")
    idx_route = find_col(header_row, COL_PUD_RTE, "PUD RTE")
    idx_ckpt = find_col(header_row, COL_ACT_CKPT, "ACT CKPT")
    if idx_ckpt is None:
        return {}, {}

    counts_global = defaultdict(int)
    by_route_raw = defaultdict(lambda: defaultdict(int))

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row = list(row)
        ckpt = (row[idx_ckpt] if idx_ckpt < len(row) else None)
        ckpt = str(ckpt).strip().upper() if ckpt else ""
        if ckpt not in OPMS_CODES:
            continue
        depot = str(row[idx_depot] or "").strip() if idx_depot < len(row) else ""
        route = str(row[idx_route] or "").strip() if idx_route < len(row) else ""
        counts_global[ckpt] += 1
        by_route_raw[(depot, route)][ckpt] += 1

    by_route = {}
    for (depot, route), cnt in by_route_raw.items():
        key = depot + "|" + route
        by_route[key] = {"depot": depot, "route": route, "counts": dict(cnt)}
    return dict(counts_global), by_route


def parse_tw(ws):
    """Sheet Export: coluna Route e '% TW Adh DL'. Retorna { routeName: value } (0-1)."""
    rows = list(ws.iter_rows(max_row=500, values_only=True))
    if not rows:
        return {}
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_route = -1
    col_tw = -1
    for i, h in enumerate(header):
        if re.search(r"route", h, re.I) and col_route < 0:
            col_route = i
        if re.search(r"%\s*TW\s*Adh\s*DL|TW Adh DL", h, re.I):
            col_tw = i
    if col_route < 0 or col_tw < 0:
        return {}
    tw_by_route = {}
    for row in rows[1:]:
        row = list(row)
        route = str(row[col_route] or "").strip() if col_route < len(row) else ""
        val = row[col_tw] if col_tw < len(row) else None
        if not route:
            continue
        num = None
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            num = float(val)
        elif val is not None and str(val).strip():
            try:
                num = float(str(val).replace(",", ".").replace(" ", ""))
            except ValueError:
                pass
        if num is not None:
            tw_by_route[route] = num
    return tw_by_route


def _normalise_subpostcode(val):
    """Converte valor (postcode ou subpostcode) para subpostcode: trim e remove últimos 2 chars se longo."""
    s = str(val or "").strip().replace(" ", "")
    if len(s) <= 2:
        return s if s else None
    return s[:-2] if len(s) > 2 else s


def _find_col(header, names):
    """Procura coluna cujo cabeçalho coincide com um dos nomes (case-insensitive)."""
    for i, h in enumerate(header):
        h = (h or "").strip().upper()
        for n in names:
            if n.upper() in h or h in n.upper():
                return i
    return -1


def load_excel_rows(path):
    """Carrega uma folha Excel (.xlsx ou .xls) como lista de linhas (cada linha = lista de valores).
    Retorna list[list]; a primeira linha pode ser cabeçalho."""
    path = os.path.normpath(os.path.expanduser(path))
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        try:
            import pandas as pd
            df = pd.read_excel(path, sheet_name=0, header=None, engine="xlrd")
            return [[_xls_cell(c) for c in row] for row in df.values.tolist()]
        except Exception as e:
            raise RuntimeError("Ficheiro .xls: instale pandas e xlrd (pip install pandas xlrd): %s" % e)
    # .xlsx ou outro
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active or wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(max_row=5000, values_only=True))
        wb.close()
        return [[str(c).strip() if c is not None else "" for c in row] for row in rows]
    except Exception as e:
        raise RuntimeError("Erro ao ler Excel: %s" % e)


def _xls_cell(c):
    """Converte célula pandas/xlrd para string para consistência."""
    if c is None or (isinstance(c, float) and c != c):  # NaN
        return ""
    if isinstance(c, float) and c == int(c):
        return str(int(c))
    return str(c).strip()


def parse_disco_from_rows(rows):
    """Mesma lógica que parse_disco(ws) mas recebe lista de linhas (list[list]).
    rows[0] pode ser cabeçalho; procura linha com 'Route' e colunas Cycle/SortWave."""
    if not rows:
        return {"am": [], "pm": [], "byRoute": {}, "raw_deliveries": []}
    header_row_idx = 0
    for i in range(min(5, len(rows))):
        row = [str(c).strip() if c is not None else "" for c in rows[i]]
        if "Route" not in row:
            continue
        if "Cycle" not in row and "SortWave" not in (row or []) and not any("Sort" in (h or "") for h in row):
            continue
        header_row_idx = i
        if any("Postal Code" in (h or "") or "Postcode" in (h or "") or "Subpostcode" in (h or "") for h in row):
            break
    header = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
    col_route = header.index("Route") if "Route" in header else -1
    col_sort = -1
    col_loop = -1
    col_depot = -1
    col_subpostcode = -1
    col_address = -1
    col_recipient = -1
    for i, h in enumerate(header):
        h = (h or "").strip()
        if "SortWave" in h or "Sort Wave" in h:
            col_sort = i
        if h == "Loop" or h == "Cycle":
            col_loop = i
        if h == "Depot" or "PUD SVC" in h or "PUD Svc" in h:
            col_depot = i
        if "Subpostcode" in h or "Sub Postcode" in h or "Sub postcode" in h:
            col_subpostcode = i
        if col_subpostcode < 0 and (h == "Postcode" or "Post code" in h or h == "Postal Code"):
            col_subpostcode = i
        if h == "Address":
            col_address = i
        if h == "Receiver" or (h == "Name" and col_address < 0):
            col_recipient = i
        elif h == "Name":
            col_recipient = i
    col_pre12 = _find_col(header, ("Pre-12", "Pre 12", "Pre12"))
    col_asr = _find_col(header, ("ASR", "Adult Signature", "Adult Signature Required"))
    col_dsr = _find_col(header, ("DSR", "Direct Signature", "Direct Signature Required"))

    def _is_flagged(r, c):
        if c < 0 or c >= len(r):
            return False
        v = r[c]
        if v is None:
            return False
        s = str(v).strip().upper()
        return s in ("1", "X", "Y", "YES", "TRUE", "S", "SIM")

    if col_route < 0:
        return {"am": [], "pm": [], "byRoute": {}, "raw_deliveries": []}
    am_set = set()
    pm_set = set()
    by_route = {}
    raw_deliveries = []
    for row in rows[header_row_idx + 1:]:
        row = list(row)
        route = str(row[col_route] if col_route < len(row) else "" or "").strip()
        if not route or route.upper() == "ROUTE":
            continue
        wave = str(row[col_sort] or "").strip().upper() if col_sort >= 0 and col_sort < len(row) else ""
        is_am = "AM" in wave or wave == "AM"
        is_pm = "PM" in wave or wave == "PM"
        if is_am:
            am_set.add(route)
        if is_pm:
            pm_set.add(route)
        loop = str(row[col_loop] or "").strip() if col_loop >= 0 and col_loop < len(row) else None
        depot = str(row[col_depot] or "").strip() if col_depot >= 0 and col_depot < len(row) else None
        sub_raw = str(row[col_subpostcode] or "").strip() if col_subpostcode >= 0 and col_subpostcode < len(row) else None
        sub = _normalise_subpostcode(sub_raw) if sub_raw else None
        if route not in by_route:
            by_route[route] = {"sortWave": "AM" if is_am else "PM", "loop": None, "depot": None, "subpostcodes": []}
        if loop:
            by_route[route]["loop"] = loop
        if depot:
            by_route[route]["depot"] = depot
        if sub and sub not in by_route[route]["subpostcodes"]:
            by_route[route]["subpostcodes"].append(sub)
        address = str(row[col_address] or "").strip() if col_address >= 0 and col_address < len(row) else ""
        recipient = str(row[col_recipient] or "").strip() if col_recipient >= 0 and col_recipient < len(row) else ""
        raw_deliveries.append({
            "depot": depot or "",
            "route": route,
            "loop": loop or "",
            "subpostcode": sub or (sub_raw or "").strip() or "",
            "address": address,
            "recipient": recipient,
            "pre12": _is_flagged(row, col_pre12),
            "asr": _is_flagged(row, col_asr),
            "dsr": _is_flagged(row, col_dsr),
        })
    for r in by_route:
        by_route[r]["subpostcodes"].sort()
    return {"am": sorted(am_set), "pm": sorted(pm_set), "byRoute": by_route, "raw_deliveries": raw_deliveries}


def merge_disco_data(data_list):
    """Junta vários resultados de parse_disco/parse_disco_from_rows numa única estrutura.
    Concatena raw_deliveries, faz união de am/pm e merge de byRoute (subpostcodes unidos)."""
    if not data_list:
        return {"am": [], "pm": [], "byRoute": {}, "raw_deliveries": []}
    am_set = set()
    pm_set = set()
    by_route = {}
    raw_deliveries = []
    for d in data_list:
        am_set.update(d.get("am") or [])
        pm_set.update(d.get("pm") or [])
        raw_deliveries.extend(d.get("raw_deliveries") or [])
        for route, info in (d.get("byRoute") or {}).items():
            if route not in by_route:
                by_route[route] = {"sortWave": info.get("sortWave", "AM"), "loop": info.get("loop"), "depot": info.get("depot"), "subpostcodes": list(info.get("subpostcodes") or [])}
            else:
                for sub in info.get("subpostcodes") or []:
                    if sub and sub not in by_route[route]["subpostcodes"]:
                        by_route[route]["subpostcodes"].append(sub)
                if info.get("loop") and not by_route[route]["loop"]:
                    by_route[route]["loop"] = info["loop"]
                if info.get("depot") and not by_route[route]["depot"]:
                    by_route[route]["depot"] = info["depot"]
    for r in by_route:
        by_route[r]["subpostcodes"].sort()
    return {"am": sorted(am_set), "pm": sorted(pm_set), "byRoute": by_route, "raw_deliveries": raw_deliveries}


def parse_disco(ws):
    """Sheet com Route e SortWave (AM/PM). Opcional: Loop, Cycle, Depot, Subpostcode/Postcode.
    Retorna { am: [routes], pm: [routes], byRoute: { route: { sortWave, loop?, depot?, subpostcodes: [] } } }."""
    rows = list(ws.iter_rows(max_row=2000, values_only=True))
    if not rows:
        return {"am": [], "pm": [], "byRoute": {}}
    header_row_idx = 0
    for i in range(min(5, len(rows))):
        row = [str(c).strip() if c is not None else "" for c in rows[i]]
        if "Route" not in row:
            continue
        if "Cycle" not in row and "SortWave" not in (row or []) and not any("Sort" in (h or "") for h in row):
            continue
        header_row_idx = i
        if any("Postal Code" in (h or "") or "Postcode" in (h or "") or "Subpostcode" in (h or "") for h in row):
            break
    header = [str(c).strip() if c is not None else "" for c in rows[header_row_idx]]
    col_route = header.index("Route") if "Route" in header else -1
    col_sort = -1
    col_loop = -1
    col_depot = -1
    col_subpostcode = -1
    col_address = -1
    col_recipient = -1
    for i, h in enumerate(header):
        h = (h or "").strip()
        if "SortWave" in h or "Sort Wave" in h:
            col_sort = i
        if h == "Loop" or h == "Cycle":
            col_loop = i
        if h == "Depot" or "PUD SVC" in h or "PUD Svc" in h:
            col_depot = i
        if "Subpostcode" in h or "Sub Postcode" in h or "Sub postcode" in h:
            col_subpostcode = i
        if col_subpostcode < 0 and (h == "Postcode" or "Post code" in h or h == "Postal Code"):
            col_subpostcode = i
        if h == "Address":
            col_address = i
        if h == "Receiver" or (h == "Name" and col_address < 0):
            col_recipient = i
        elif h == "Name":
            col_recipient = i
    col_pre12 = _find_col(header, ("Pre-12", "Pre 12", "Pre12"))
    col_asr = _find_col(header, ("ASR", "Adult Signature", "Adult Signature Required"))
    col_dsr = _find_col(header, ("DSR", "Direct Signature", "Direct Signature Required"))

    def _is_flagged(r, c):
        if c < 0 or c >= len(r):
            return False
        v = r[c]
        if v is None:
            return False
        s = str(v).strip().upper()
        return s in ("1", "X", "Y", "YES", "TRUE", "S", "SIM")

    if col_route < 0:
        return {"am": [], "pm": [], "byRoute": {}, "raw_deliveries": []}
    am_set = set()
    pm_set = set()
    by_route = {}
    raw_deliveries = []
    for row in rows[header_row_idx + 1:]:
        row = list(row)
        route = str(row[col_route] or "").strip() if col_route < len(row) else ""
        if not route or route.upper() == "ROUTE":
            continue
        wave = str(row[col_sort] or "").strip().upper() if col_sort >= 0 and col_sort < len(row) else ""
        is_am = "AM" in wave or wave == "AM"
        is_pm = "PM" in wave or wave == "PM"
        if is_am:
            am_set.add(route)
        if is_pm:
            pm_set.add(route)
        loop = str(row[col_loop] or "").strip() if col_loop >= 0 and col_loop < len(row) else None
        depot = str(row[col_depot] or "").strip() if col_depot >= 0 and col_depot < len(row) else None
        sub_raw = str(row[col_subpostcode] or "").strip() if col_subpostcode >= 0 and col_subpostcode < len(row) else None
        sub = _normalise_subpostcode(sub_raw) if sub_raw else None
        if route not in by_route:
            by_route[route] = {"sortWave": "AM" if is_am else "PM", "loop": None, "depot": None, "subpostcodes": []}
        if loop:
            by_route[route]["loop"] = loop
        if depot:
            by_route[route]["depot"] = depot
        if sub and sub not in by_route[route]["subpostcodes"]:
            by_route[route]["subpostcodes"].append(sub)
        address = str(row[col_address] or "").strip() if col_address >= 0 and col_address < len(row) else ""
        recipient = str(row[col_recipient] or "").strip() if col_recipient >= 0 and col_recipient < len(row) else ""
        raw_deliveries.append({
            "depot": depot or "",
            "route": route,
            "loop": loop or "",
            "subpostcode": sub or (sub_raw or "").strip() or "",
            "address": address,
            "recipient": recipient,
            "pre12": _is_flagged(row, col_pre12),
            "asr": _is_flagged(row, col_asr),
            "dsr": _is_flagged(row, col_dsr),
        })
    for r in by_route:
        by_route[r]["subpostcodes"].sort()
    return {"am": sorted(am_set), "pm": sorted(pm_set), "byRoute": by_route, "raw_deliveries": raw_deliveries}


def build_loops_from_raw(raw_deliveries):
    """Agrupa entregas por loop; cada loop tem totalDeliveries e lista de rotas com quantidade.
    Retorna dict: { loopId: { totalDeliveries: N, routes: [ { route: "R1", deliveries: n }, ... ] } }."""
    loops = {}
    for r in raw_deliveries or []:
        loop_id = (r.get("loop") or "").strip() or "—"
        route = (r.get("route") or "").strip()
        if not loop_id:
            loop_id = "—"
        if loop_id not in loops:
            loops[loop_id] = {"totalDeliveries": 0, "routes": {}}
        loops[loop_id]["totalDeliveries"] += 1
        if route:
            loops[loop_id]["routes"][route] = loops[loop_id]["routes"].get(route, 0) + 1
    # Converter routes de dict para lista ordenada
    out = {}
    for lid, data in loops.items():
        out[lid] = {
            "totalDeliveries": data["totalDeliveries"],
            "routes": [{"route": r, "deliveries": c} for r, c in sorted(data["routes"].items())],
        }
    return out


def disco_to_dashboard_json(disco_data):
    """Converte o resultado de parse_disco() para o formato do dashboard (disco.json).
    Retorna dict com: deliveries, depot, loop, routes, source, loops.
    Se existir raw_deliveries (uma linha por parada), usa-o; senão constrói a partir de by_route."""
    by_route = disco_data.get("byRoute") or {}
    raw = disco_data.get("raw_deliveries") or []
    loops = build_loops_from_raw(raw)
    if raw:
        deliveries = [
            {
                "depot": r.get("depot", ""),
                "route": r.get("route", ""),
                "loop": r.get("loop", ""),
                "subpostcode": r.get("subpostcode", ""),
                "address": r.get("address", ""),
                "recipient": r.get("recipient", ""),
                "pre12": bool(r.get("pre12")),
                "asr": bool(r.get("asr")),
                "dsr": bool(r.get("dsr")),
            }
            for r in raw
        ]
        depot_seen = next((r.get("depot") or "" for r in raw if r.get("depot")), "")
        loop_seen = (by_route.get(raw[0].get("route", "")) or {}).get("loop", "") if raw else ""
    else:
        deliveries = []
        depot_seen = None
        loop_seen = None
        for route, info in by_route.items():
            d = info.get("depot") or ""
            l = info.get("loop") or ""
            if d and depot_seen is None:
                depot_seen = d
            if l and loop_seen is None:
                loop_seen = l
            for sub in info.get("subpostcodes") or []:
                deliveries.append({
                    "depot": d,
                    "route": route,
                    "subpostcode": sub,
                    "address": "",
                    "recipient": ""
                })
    routes = sorted(set(disco_data.get("am") or []) | set(disco_data.get("pm") or []))
    if not routes:
        routes = sorted(by_route.keys())
    if not depot_seen and by_route:
        depot_seen = next((info.get("depot") or "" for info in by_route.values() if info.get("depot")), "")
    if not loop_seen and by_route:
        loop_seen = next((info.get("loop") or "" for info in by_route.values() if info.get("loop")), "")
    return {
        "source": "DISCO_RouteData.xlsx",
        "depot": depot_seen or "",
        "loop": loop_seen or "",
        "routes": routes,
        "deliveries": deliveries,
        "loops": loops,
    }


def build_by_route_from_deliveries(deliveries):
    """Constrói byRoute no formato esperado pelo dashboard (route -> { subpostcodes: [] })."""
    by_route = {}
    for row in deliveries:
        route = (row.get("route") or "").strip()
        sub = row.get("subpostcode")
        if not route:
            continue
        if route not in by_route:
            by_route[route] = {}
        if sub:
            by_route[route][str(sub).strip()] = True
    return {r: {"subpostcodes": sorted(by_route[r].keys())} for r in by_route}


def write_disco_assets(dashboard_disco):
    """Escreve assets/disco.json e assets/disco-data.js. Retorna True em sucesso."""
    deliveries = dashboard_disco.get("deliveries") or []
    by_route = build_by_route_from_deliveries(deliveries)
    payload = {**dashboard_disco, "byRoute": by_route, "loops": dashboard_disco.get("loops") or {}}
    try:
        os.makedirs(ASSETS_DIR, exist_ok=True)
        disco_json_path = os.path.join(ASSETS_DIR, "disco.json")
        with open(disco_json_path, "w", encoding="utf-8") as f:
            json.dump(dashboard_disco, f, indent=2, ensure_ascii=False)
        print("Escrito:", disco_json_path, "(%d deliveries)" % len(deliveries))
        disco_js_path = os.path.join(ASSETS_DIR, "disco-data.js")
        with open(disco_js_path, "w", encoding="utf-8") as f:
            f.write("/* Gerado por scripts/import_excel_to_data.py */\nwindow.DISCO_DATA = ")
            json.dump(payload, f, ensure_ascii=False)
            f.write(";\n")
        print("Escrito:", disco_js_path)
        return True
    except OSError as e:
        print("Aviso: não foi possível escrever assets do Disco:", e, file=sys.stderr)
        return False


def main():
    default_downloads = os.path.expanduser("~/Downloads")
    path_ba = os.path.join(default_downloads, "BA Daily Figures 3.xlsx")
    path_tw = os.path.join(default_downloads, "TW 300126.xlsx")
    path_disco = os.path.join(default_downloads, "DISCO_RouteData_20260202_064654.xlsx")

    if len(sys.argv) >= 4:
        a, b, c = sys.argv[1], sys.argv[2], sys.argv[3]
        if _norm_path(a):
            path_ba = _norm_path(a)
        if _norm_path(b):
            path_tw = _norm_path(b)
        if _norm_path(c):
            path_disco = _norm_path(c)
    elif len(sys.argv) == 2 and _norm_path(sys.argv[1]):
        path_disco = _norm_path(sys.argv[1])

    opms_date = DEFAULT_OPMS_DATE
    disco_data = {"am": [], "pm": [], "byRoute": {}}
    opms_counts = {}
    opms_by_route = {}
    tw_by_route = {}

    # Disco
    if path_disco and os.path.isfile(path_disco):
        try:
            wb = openpyxl.load_workbook(path_disco, read_only=True, data_only=True)
            for name in wb.sheetnames:
                disco_data = parse_disco(wb[name])
                if disco_data["am"] or disco_data["pm"] or disco_data.get("byRoute"):
                    break
            wb.close()
        except Exception as e:
            print("Erro ao ler ficheiro Disco:", e, file=sys.stderr)
        else:
            write_disco_assets(disco_to_dashboard_json(disco_data))
    else:
        if path_disco:
            print("Ficheiro Disco não encontrado:", path_disco, file=sys.stderr)
        else:
            print("Caminho Disco não indicado.", file=sys.stderr)

    # Time Window
    if path_tw and os.path.isfile(path_tw):
        try:
            wb = openpyxl.load_workbook(path_tw, read_only=True, data_only=True)
            for name in wb.sheetnames:
                tw_by_route = parse_tw(wb[name])
                if tw_by_route:
                    break
            wb.close()
        except Exception as e:
            print("Erro ao ler ficheiro TW:", e, file=sys.stderr)
    elif path_tw:
        print("Ficheiro TW não encontrado:", path_tw, file=sys.stderr)

    # OPMS (BA Daily Figures)
    if path_ba and os.path.isfile(path_ba):
        try:
            wb = openpyxl.load_workbook(path_ba, read_only=True, data_only=True)
            for name in wb.sheetnames:
                if OPMS_SHEET_NAME.lower() in name.lower():
                    opms_counts, opms_by_route = parse_opms(wb[name])
                    break
            if not opms_counts and not opms_by_route:
                for name in wb.sheetnames:
                    opms_counts, opms_by_route = parse_opms(wb[name])
                    if opms_counts or opms_by_route:
                        break
            wb.close()
        except Exception as e:
            print("Erro ao ler ficheiro BA:", e, file=sys.stderr)
    elif path_ba:
        print("Ficheiro BA não encontrado:", path_ba, file=sys.stderr)

    # dhl-embedded-data.js
    out_path = os.path.join(REPO_ROOT, "dhl-embedded-data.js")
    opms_for_date = {
        "counts": opms_counts,
        "byRoute": opms_by_route,
        "twByRoute": tw_by_route,
        "incomeByRouteFromExcel": {}
    }
    js_content = """/**
 * Dados importados dos ficheiros Excel (DISCO, OPMS, Time Window).
 * Gerado por scripts/import_excel_to_data.py
 * - DISCO_RouteData: rotas AM/PM
 * - BA Daily Figures 3: OPMS (counts, byRoute)
 * - TW 300126: Time Window (% TW Adh DL) por rota
 */
(function () {
  'use strict';
  window.DISCO_DATA = """
    js_content += json.dumps(disco_data, indent=2, ensure_ascii=False)
    js_content += """;

  window.OPMS_EMBEDDED_DATA = """
    js_content += json.dumps({opms_date: opms_for_date}, indent=2, ensure_ascii=False)
    js_content += """;
})();
"""
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(js_content)
        print("Escrito:", out_path)
        print("  DISCO: am=%d, pm=%d rotas" % (len(disco_data["am"]), len(disco_data["pm"])))
        print("  OPMS: %s counts, %d byRoute" % (opms_date, len(opms_by_route)))
        print("  TW: %d rotas" % len(tw_by_route))
    except OSError as e:
        print("Erro ao escrever dhl-embedded-data.js:", e, file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
