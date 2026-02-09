#!/usr/bin/env python3
"""
OPMS (Last Day Operations) – Processamento de dados de produção.

Mapeamento:
  PUD Svc Area  → Depot
  PUD Rte       → Route
  Courier id    → Courier (Driver)
  Act Ckpt Code → Códigos de confirmação (apenas: BA, CA, CM, FP, HN, NR, OK, PU, RD)

Círculos de métricas:
  Deliveries = PU + HN + OK
  AFD        = BA + CA + CM + FP + NR + RD

Income = (PU + HN + OK) × rate do loop (banda).

Uso:
  python opms_process.py "path/to/BA Daily Figures 3.xlsx" [--rate RATE] [--sheet NAME]
"""

import argparse
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Requer openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Códigos Act Ckpt que devem ser processados e contados
OPMS_CODES = frozenset({"BA", "CA", "CM", "FP", "HN", "NR", "OK", "PU", "RD"})
DELIVERIES_CODES = frozenset({"PU", "HN", "OK"})  # Círculo Deliveries
# AFD = todos os outros em OPMS_CODES

# Nomes de colunas no Excel (sheet OPMS - RDT ROADMAP Stop File)
COL_PUD_SVC = "PUD Svc Area"   # Depot
COL_PUD_RTE = "PUD Rte"        # Route
COL_COURIER = "Courier id"    # Courier (Driver)
COL_ACT_CKPT = "Act Ckpt Code"

DEFAULT_OPMS_SHEET = "OPMS - RDT ROADMAP Stop File"


def find_header_row(ws):
    """Encontra a linha do header (primeira linha que contém 'Act Ckpt Code')."""
    for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        row_str = [str(c).strip() if c is not None else "" for c in row]
        if COL_ACT_CKPT in row_str:
            return i, row_str
        if any("Act Ckpt Code" in (s or "") for s in row):
            return i, row_str
    return None, []


def parse_opms_sheet(ws):
    """
    Lê a sheet OPMS e devolve:
    - rows: lista de dict com depot, route, courier, act_ckpt_code
    - counts_global: { code: total }
    - by_route: { (depot, route): { code: count } }
    """
    header_row_idx, header_row = find_header_row(ws)
    if not header_row:
        return None, None, None

    def find_col(name, fallback_sub=None):
        for i, h in enumerate(header_row):
            if h == name:
                return i
            if fallback_sub and (h or "").strip() and fallback_sub in (h or "").upper():
                return i
        return None

    idx_depot = find_col(COL_PUD_SVC, "PUD SVC")
    idx_route = find_col(COL_PUD_RTE, "PUD RTE")
    idx_courier = find_col(COL_COURIER, "COURIER")
    idx_ckpt = find_col(COL_ACT_CKPT, "ACT CKPT")

    if idx_ckpt is None:
        return None, None, None

    rows = []
    counts_global = defaultdict(int)
    by_route = defaultdict(lambda: defaultdict(int))

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row = list(row)
        ckpt = (row[idx_ckpt] if idx_ckpt is not None and idx_ckpt < len(row) else None)
        ckpt = str(ckpt).strip().upper() if ckpt is not None else ""
        if ckpt not in OPMS_CODES:
            continue

        depot = str(row[idx_depot] or "").strip() if idx_depot is not None and idx_depot < len(row) else ""
        route = str(row[idx_route] or "").strip() if idx_route is not None and idx_route < len(row) else ""
        courier = str(row[idx_courier] or "").strip() if idx_courier is not None and idx_courier < len(row) else ""

        rows.append({
            "depot": depot,
            "route": route,
            "courier": courier,
            "act_ckpt_code": ckpt,
        })
        counts_global[ckpt] += 1
        by_route[(depot, route)][ckpt] += 1

    return rows, dict(counts_global), {k: dict(v) for k, v in by_route.items()}


def compute_metrics(counts):
    """A partir de um dict code -> count, devolve deliveries, afd e income (income precisa de rate)."""
    deliveries = sum(counts.get(c, 0) for c in DELIVERIES_CODES)
    afd = sum(counts.get(c, 0) for c in OPMS_CODES if c not in DELIVERIES_CODES)
    return deliveries, afd


def main():
    ap = argparse.ArgumentParser(description="Processar ficheiro OPMS (Last Day Operations)")
    ap.add_argument("excel_path", help="Caminho para o ficheiro Excel (ex: BA Daily Figures 3.xlsx)")
    ap.add_argument("--rate", type=float, default=2.90, help="Rate (banda) por entrega para cálculo do Income (default: 2.90)")
    ap.add_argument("--sheet", default=DEFAULT_OPMS_SHEET, help="Nome da sheet OPMS (default: %s)" % DEFAULT_OPMS_SHEET)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.excel_path, read_only=True, data_only=True)
    sheet = None
    for name in wb.sheetnames:
        if args.sheet.lower() in name.lower() or name.strip() == args.sheet:
            sheet = wb[name]
            break
    if sheet is None:
        # Tentar primeira sheet que tenha header OPMS
        for name in wb.sheetnames:
            ws = wb[name]
            hr_idx, hr = find_header_row(ws)
            if hr and COL_ACT_CKPT in (hr or []):
                sheet = ws
                break
        if sheet is None:
            print("Sheet OPMS não encontrada. Sheets disponíveis:", wb.sheetnames, file=sys.stderr)
            wb.close()
            sys.exit(1)

    rows, counts_global, by_route = parse_opms_sheet(sheet)
    wb.close()

    if counts_global is None:
        print("Não foi possível ler dados OPMS (header não encontrado).", file=sys.stderr)
        sys.exit(1)

    # Counts por código (ordem fixa)
    print("=== Counts por código (Act Ckpt Code) ===\n")
    for code in ["BA", "CA", "CM", "FP", "HN", "NR", "OK", "PU", "RD"]:
        c = counts_global.get(code, 0)
        print("  %s: %d" % (code, c))

    deliveries, afd = compute_metrics(counts_global)
    income = deliveries * args.rate

    print("\n=== Métricas (círculos) ===")
    print("  Deliveries (PU + HN + OK): %d" % deliveries)
    print("  AFD (BA + CA + CM + FP + NR + RD): %d" % afd)
    print("\n=== Income ===")
    print("  (Deliveries × rate) = %d × %.2f = %.2f" % (deliveries, args.rate, income))

    # Por rota (resumo)
    if by_route:
        print("\n=== Por Depot / Route ===")
        for (depot, route), counts in sorted(by_route.items(), key=lambda x: (x[0][0], x[0][1])):
            d, a = compute_metrics(counts)
            inc = d * args.rate
            print("  %s / %s  →  Deliveries: %d, AFD: %d, Income: %.2f" % (depot, route, d, a, inc))

    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
